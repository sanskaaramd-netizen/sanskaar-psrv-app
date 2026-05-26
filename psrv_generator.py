from openpyxl import load_workbook
from openpyxl.styles import Alignment


def fill_boxes(ws, value, row, start_col):

    value = str(value)

    for i, char in enumerate(value):
        ws.cell(
            row=row,
            column=start_col + i
        ).value = char


def generate_psrv(
    invoice_data,
    eway_data,
    template_path,
    output_path
):

    wb = load_workbook(template_path)
    ws = wb.active

    # ==========================
    # GSTIN
    # ==========================
    ws["AD6"] = invoice_data.get("gstin", "")

    # ==========================
    # E-Way Bill Number
    # ==========================
    ws["AD7"] = str(
        eway_data.get("eway_bill_no", "")
    )

    # ==========================
    # Invoice Number
    # ==========================
    invoice_no = invoice_data.get(
        "invoice_number",
        ""
    )

    invoice_no = invoice_no.replace(
        "INV-SA-",
        ""
    )

    invoice_no = invoice_no.replace("/", "")
    invoice_no = invoice_no.replace("-", "")

    fill_boxes(
        ws,
        invoice_no,
        12,
        7
    )  # G12

    # ==========================
    # PO Number
    # ==========================
    po_no = invoice_data.get(
        "po_number",
        ""
    )

    fill_boxes(
        ws,
        po_no,
        13,
        7
    )  # G13

    # ==========================
    # Invoice Date
    # ==========================
    invoice_date = invoice_data.get(
        "invoice_date",
        ""
    )

    if invoice_date:

        dd, mm, yyyy = invoice_date.split("/")

        # DD
        ws["AA12"] = dd

        # MM
        ws["AB12"] = mm[0]
        ws["AC12"] = mm[1]

        # YYYY
        ws["AD12"] = yyyy[:2]
        ws["AE12"] = yyyy[2:]

    # ==========================
    # Total Invoice Value
    # ==========================
    ws["AJ12"] = invoice_data.get(
        "grand_total",
        ""
    )

    # ==========================
    # Materials
    # ==========================
    materials = invoice_data.get(
        "materials",
        []
    )

    for index, item in enumerate(materials):

        row = 16 + (index * 2)

        # ==========================
        # Material Code
        # ==========================
        fill_boxes(
            ws,
            item.get("code", ""),
            row,
            2
        )  # Column B

        # ==========================
        # Description
        # ==========================
        description = item.get(
            "description",
            ""
        )

        # Optional shortening
        if len(description) > 70:
            description = description[:70]

        desc_cell = ws[f"M{row}"]

        desc_cell.value = description

        desc_cell.alignment = Alignment(
            wrap_text=True,
            vertical="center"
        )

        # ==========================
        # Quantity
        # ==========================
        ws[f"T{row}"] = item.get(
            "qty",
            ""
        )

        # ==========================
        # HSN Code
        # ==========================
        ws[f"AA{row}"] = item.get(
            "hsn",
            ""
        )

        # ==========================
        # Basic Value
        # ==========================
        ws[f"AE{row}"] = item.get(
            "amount",
            ""
        )

        # ==========================
        # CGST Amount
        # ==========================
        ws[f"AI{row}"] = item.get(
            "cgst_amount",
            ""
        )

        # ==========================
        # SGST Amount
        # ==========================
        ws[f"AJ{row}"] = item.get(
            "sgst_amount",
            ""
        )

    wb.save(output_path)

    print("PSRV Generated Successfully")
